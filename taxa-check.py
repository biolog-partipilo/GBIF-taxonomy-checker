import requests
from openpyxl import load_workbook
from tqdm import tqdm
import time

# Excel file path
FILE_PATH = r"YOUR-PATH-HERE-NAMEFILE.xlsx"

# Save interval - default = 25
SAVE_EVERY = 25  

# GBIF API endpoint
GBIF_URL = "https://api.gbif.org/v1/species/match"

# Create a reusable HTTP session
session = requests.Session()

def check_species(name):
    """
    Query GBIF and return:
    - status
    - accepted name
    - author/year (if available)
    """
 
    try:
        response = session.get(
            GBIF_URL,
            params={"name": name},
            timeout=6
        )

        if response.status_code != 200:
            print(f"❌ API error for {name}: {response.status_code}")
            return "error", "", ""

        data = response.json()

        status = data.get("status")
        match_type = data.get("matchType")

        scientific = data.get("scientificName", "")
        accepted = data.get("acceptedScientificName", "")
        auth = data.get("authorship", "")

        # CASE 1: accepted name
        if status == "ACCEPTED":
            return "valid", scientific, auth

        # CASE 2: synonym → do backbone check
        if status == "SYNONYM":
            key = data.get("usageKey")

            if key:
                full = session.get(
                    f"https://api.gbif.org/v1/species/{key}",
                    timeout=6
                ).json()

                accepted_name = full.get("canonicalName") or full.get("scientificName", "")
                authorship = full.get("authorship", "")

                return "syn", accepted_name, authorship

            return "syn", accepted, auth

        # CASE 3: name not found
        if match_type in ["NONE", "HIGHERRANK"]:
            print(f"⚠️ NOT FOUND: {name}")
            return "missing", "", ""

        return "error", "", ""

    except Exception as e:
        print(f"❌ Exception for {name}: {e}")
        return "error", "", ""

def main():
    wb = load_workbook(FILE_PATH)
    ws = wb.active

    total_rows = ws.max_row
# show work in progress
    with tqdm(
        total=total_rows - 2,
        desc="Processing",
        unit="species",
        ncols=80,
        dynamic_ncols=False,
        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{percentage:3.0f}%]"
    ) as pbar:

        for row in range(3, total_rows + 1):
            name = ws[f"D{row}"].value

            if not name:
                pbar.update(1)
                continue

            tqdm.write(f"checking {row}: {name}")

            status, updated_name, author = check_species(name)

            # Write results
            ws[f"I{row}"] = status
            ws[f"J{row}"] = updated_name
            ws[f"K{row}"] = author

            # Clean author (remove parentheses)
            if ws[f"K{row}"].value:
                ws[f"K{row}"].value = ws[f"K{row}"].value.split("(")[0].strip()

            # Save periodically
            if row % SAVE_EVERY == 0:
                wb.save(FILE_PATH)
                print(f"💾 saved up to row {row}")

            time.sleep(0.05)
            pbar.update(1)
# job end
    wb.save(FILE_PATH)
    print("✅ completed!")


if __name__ == "__main__":
    main()
